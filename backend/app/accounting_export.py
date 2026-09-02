from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from typing import Any, DefaultDict, Dict, Iterable, List, Tuple

import httpx
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

from .models import ProgressEvent, RunRequest
from .service import _build_headers, _handle_response, _refresh_access_token, _resolve_backoffice


KITCHEN_REUNION_BRANCHES = {
    "7302bb40a44f4efa816524e20a8f186a": "Kitchen Reunion - Welle 7",
    "eab690581d2a4ca0a793c212ff3ab58f": "Shoppyland, Schoenbuehl",
    "796a0ca675b44eafa46649a40c59bfd5": "Shoppy Tivoli, Thai Wok",
    "a3b12f0836f14011a28c81a7057cf4f9": "Shoppy Tivoli, Guapo Mexicano",
    "ae14f5980672453192b492ea7ed42092": "Mall Of Switzerland",
    "a863fcd8c4034bcdb843d9354ac2f0b4": "Stueckipark",
}

LAAX_BRANCHES = {
    "5d2a98c8dc504b3abebd425dc797e59c": "Camino MA",
    "9f66bc2e05d54fb6bba79ec20f73d64f": "Piazza Cafedeli",
    "e6e1779fcfe14b41993768a8e3f31536": "Segneshütte",
    "2e6398772566478e91f35f46c3561675": "Ella",
    "4a16b0f8e78c4a59b17e21fc857180ff": "Restaurant Ikigai",
    "8b998652fb4c475e8b8e7795b6aac013": "Burgers",
    "9195daaea11f4e46b5dd73caab9ba300": "Camino Take Away",
}

PIZZA_NATION_BRANCHES = {
    "768255ee372b42238e1e39aa3a6832da": "Pizza Nation Binz",
    "d41818219d6f44ed86daf77d65414c7f": "Pizza Nation Rosengasse",
}

VAPIANO_BRANCHES = {
    "b4ead0e87ce640fabad9ab5af4495e7b": "Vapiano",
}

BSC_YOUNG_BOYS_BRANCHES = {
    "def9d93fb54041d7a706d942cf443eb6": "BOX 6 - Pizza",
    "6829501dee31430ea7ad0e9fb8262939": "BOX 7 - Pommes",
    "d98039abca7944b3a62fe9837a560934": "BOX 8 - Grill",
    "e66ee55df56b410fa17175fc4296d291": "BOX 9 - Döner",
}

DADDY_FOOD_BRANCHES = {
    "0092f94414d0472397042367d60d36ca": "Daddy Food",
}

CRUSTOPIA_BRANCHES = {
    "c7011d48887d44948fd973e7bf119e71": "Crustopia KLG",
}

BURGERMEISTER_BRANCHES = {
    "1d470650037549598bfcd7030248ca20": "BM Kaserne",
    "e35ed51974f8463ab2a2f408fcd72151": "BM Altstetten",
    "b7443793591d4e4c93b58583f7c9083b": "BM Gerbergasse",
    "27eb7c620f774d0d81b0f93a6baad65c": "BM Enge",
    "d9c34c58a92e4872bd098283079d85a6": "BM Spisertor",
    "091b0b59c82c4d8397f04a7947c0dc38": "BM Claraplatz",
    "963321fc85c34f9ea55e0c2215a290a6": "BM Eschenvorstadt",
    "6e93fdc58b804263a0022e138d7785ce": "BM Escherwyss",
    "4a4c04f0556648809e8618b8afc0f4fb": "BM Langstrasse",
    "9b7ce726a94d4eac9e7fac8e891f2c7f": "BM Limmatplatz",
    "976ebb7656be460086d78f2fa1ce9772": "BM Oberdorf",
    "46cfdc648fb84c0bad5b40fe5e6f0bee": "BM Oerlikon",
    "142f98a760554108aba54c34c558fde7": "BM Winterthur",
}

# Combined lookup across all orgs that support the accounting export.
BRANCH_LABELS = {**KITCHEN_REUNION_BRANCHES, **LAAX_BRANCHES, **PIZZA_NATION_BRANCHES, **VAPIANO_BRANCHES, **BSC_YOUNG_BOYS_BRANCHES, **DADDY_FOOD_BRANCHES, **CRUSTOPIA_BRANCHES, **BURGERMEISTER_BRANCHES}

# Maps an organization id (as sent in the run request) to a human-readable
# label used in the summary sheet title.
ORG_LABELS = {
    "04675bce-b515-4eea-89d0-b74cf0db1bd5": "Kitchen Reunion",
    "d2723a46-6d95-4a46-ac92-8fbb64cba2ef": "LAAX",
    "ba38dceb-4221-455a-8fbe-610c63aa159f": "Pizza Nation",
    "1362514e-5f01-47e2-ac02-4d42e56abe2b": "Vapiano",
    "95e43289-85d5-49dd-8204-afe17c51a3ef": "BSC Young Boys",
    "611ec831-3e8d-44bb-8c16-fbc2bb1793ea": "Daddy Food",
    "83930a21-48d8-485c-9bcd-6f18b28806c6": "Crustopia KLG",
}

ACCOUNTING_QUERY_TEMPLATE = """query findData {
  PlOrders(Model: {PageNumber: {{PAGE_NUMBER}}, Filter: "{ 'OrganizationId': '{{ORG_ID}}', 'BranchUUID': '{{BRANCH_UUID}}' ,'CreateDate': {'$lte': ISODate('{{END_DATE}}'), '$gte': ISODate('{{START_DATE}}') }, }", Sort: "{CreateDate: -1}", PageSize: {{PAGE_SIZE}}}) {
    Data {
      ItemId
      CreateDate
      OrderNumber
      BranchUUID
      Device {
        DeviceType
      }
      TotalAmount
      SubTotal
      DiscountAmount
      TaxAmount
      TipAmount
      DeliveryCost
      PaymentMethod
      PaymentReferenceId
      OrderStatus
      OrderType
      OrderProducts {
        Name
        ProductId
        ProductVariationName
        UnitPrice
        Quantity
        CategoryName
        TaxId
        TaxRate
        DiscountAmount
        OrderProductModifiers {
          Name
          Price
          Quantity
        }
        RefundQuantity
      }
    }
    TotalCount
    Success
    ErrorMessage
  }
}"""

MONEY_QUANT = Decimal("0.01")


@dataclass
class SummaryBucket:
    gross_total: Decimal = Decimal("0")
    net_total: Decimal = Decimal("0")
    discounts_total: Decimal = Decimal("0")
    tips_total: Decimal = Decimal("0")
    tax_due_81: Decimal = Decimal("0")
    tax_due_26: Decimal = Decimal("0")
    revenue_81: Decimal = Decimal("0")
    revenue_26: Decimal = Decimal("0")
    # Stornierungen / Gutschriften (refunds), computed as unit price * refunded
    # quantity per product line. Tracked separately so the summary can show the
    # full breakdown before they are deducted from the totals.
    refunds_gross: Decimal = Decimal("0")
    refunds_net: Decimal = Decimal("0")
    refund_tax_81: Decimal = Decimal("0")
    refund_tax_26: Decimal = Decimal("0")
    refund_revenue_81: Decimal = Decimal("0")
    refund_revenue_26: Decimal = Decimal("0")
    refund_count: int = 0
    payment_methods: DefaultDict[str, Decimal] | None = None

    def __post_init__(self) -> None:
        if self.payment_methods is None:
            self.payment_methods = defaultdict(lambda: Decimal("0"))


def _normalize_branch_id(value: str) -> str:
    return value.strip().lower().replace("-", "")


def _branch_label(branch_id: str) -> str:
    return BRANCH_LABELS.get(_normalize_branch_id(branch_id), branch_id)


def _org_label(org_id: str | None) -> str:
    return ORG_LABELS.get((org_id or "").strip().lower(), "OMK")


def _decimal(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _money(value: Decimal) -> float:
    return float(value.quantize(MONEY_QUANT, rounding=ROUND_HALF_UP))


def _bucket_for_rate(rate: Decimal) -> str | None:
    rounded = rate.quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
    if rounded == Decimal("8.1"):
        return "81"
    if rounded == Decimal("2.6"):
        return "26"
    return None


def _modifier_total(product: Dict[str, Any]) -> Decimal:
    modifiers = product.get("OrderProductModifiers") or []
    if not isinstance(modifiers, list):
        return Decimal("0")

    total = Decimal("0")
    for modifier in modifiers:
        if not isinstance(modifier, dict):
            continue
        total += _decimal(modifier.get("Price")) * _decimal(modifier.get("Quantity"))
    return total


def _line_amounts(product: Dict[str, Any]) -> Tuple[Decimal, Decimal, Decimal, Decimal]:
    unit_price = _decimal(product.get("UnitPrice"))
    quantity = _decimal(product.get("Quantity"))
    discount = _decimal(product.get("DiscountAmount"))
    rate = _decimal(product.get("TaxRate"))
    modifiers_total = _modifier_total(product)

    gross = (unit_price * quantity) + modifiers_total - discount
    if gross < 0:
        gross = Decimal("0")

    if rate <= 0:
        net = gross
        tax = Decimal("0")
    else:
        divisor = Decimal("1") + (rate / Decimal("100"))
        net = gross / divisor
        tax = gross - net

    return gross, net, tax, rate


def _refunded_quantity(product: Dict[str, Any]) -> Decimal:
    qty = _decimal(product.get("RefundQuantity"))
    return qty if qty > 0 else Decimal("0")


def _refund_amounts(product: Dict[str, Any]) -> Tuple[Decimal, Decimal, Decimal, Decimal, Decimal]:
    """Refund (Storno/Gutschrift) amounts for a product line.

    Per accounting requirement the refund value is unit price * refunded
    quantity (gross, incl. VAT). Returns (gross, net, tax, rate, refunded_qty).
    """
    unit_price = _decimal(product.get("UnitPrice"))
    refunded_qty = _refunded_quantity(product)
    rate = _decimal(product.get("TaxRate"))

    refund_gross = unit_price * refunded_qty
    if refund_gross < 0:
        refund_gross = Decimal("0")

    if rate <= 0:
        refund_net = refund_gross
        refund_tax = Decimal("0")
    else:
        divisor = Decimal("1") + (rate / Decimal("100"))
        refund_net = refund_gross / divisor
        refund_tax = refund_gross - refund_net

    return refund_gross, refund_net, refund_tax, rate, refunded_qty


def _build_payload(req: RunRequest, branch_uuid: str, page_number: int) -> Dict[str, Any]:
    replacements = {
        "{{ORG_ID}}": req.orgId or "",
        "{{BRANCH_UUID}}": branch_uuid,
        "{{PAGE_NUMBER}}": str(page_number),
        "{{PAGE_SIZE}}": str(req.pageSize),
        "{{START_DATE}}": req.startDate or "",
        "{{END_DATE}}": req.endDate or "",
    }
    query = ACCOUNTING_QUERY_TEMPLATE
    for key, value in replacements.items():
        query = query.replace(key, value)
    return {"operationName": "", "variables": {}, "query": query}


def _parse_orders(data: Dict[str, Any]) -> List[Dict[str, Any]]:
    payload = data.get("data") or {}
    pl_orders = payload.get("PlOrders")
    if not isinstance(pl_orders, dict):
        raise RuntimeError("Could not find PlOrders in response")
    if pl_orders.get("Success") is False:
        raise RuntimeError(str(pl_orders.get("ErrorMessage") or "Backend returned Success=false"))
    items = pl_orders.get("Data") or []
    if not isinstance(items, list):
        raise RuntimeError("PlOrders.Data is not a list")
    return items


def fetch_accounting_orders(req: RunRequest) -> Tuple[List[Dict[str, Any]], List[ProgressEvent], List[str]]:
    timeout = httpx.Timeout(req.timeoutSeconds)
    backoffice = _resolve_backoffice(req.backofficeId)
    headers = _build_headers(req.cookie, req.origin, req.referer, backoffice)
    events: List[ProgressEvent] = []
    errors: List[str] = []
    all_orders: List[Dict[str, Any]] = []

    if not req.branchUuids:
        return all_orders, events, errors

    with httpx.Client(timeout=timeout, trust_env=False) as client:
        for branch in req.branchUuids:
            for page in range(1, req.maxPages + 1):
                try:
                    if req.cookie:
                        refreshed_cookie = _refresh_access_token(req.cookie, client)
                        if refreshed_cookie != req.cookie:
                            req.cookie = refreshed_cookie
                            headers["cookie"] = refreshed_cookie

                    payload = _build_payload(req, branch, page)
                    resp = client.post(backoffice["graphql_url"], headers=headers, json=payload)
                    data = _handle_response(resp)
                    orders = _parse_orders(data)
                    for o in orders:
                        o["_queried_branch"] = branch
                    all_orders.extend(orders)
                    events.append(ProgressEvent(branch=branch, page=page, status="ok"))

                    if len(orders) < req.pageSize:
                        break
                except Exception as exc:
                    errors.append(f"branch={branch}, page={page}, error={exc}")
                    events.append(ProgressEvent(branch=branch, page=page, status="error", message=str(exc)))
                    break

    return all_orders, events, errors


# Lower rank = better status to keep when deduplicating retries
_STATUS_RANK: Dict[str, int] = {"Processed": 0, "Completed": 0}


def _deduplicate_retry_orders(orders: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Remove duplicate orders created by the backoffice retry mechanism.

    When a Mobile App order fails, backoffice can resend it — which creates a
    *new* order (new ItemId) but with the same BranchUUID + OrderNumber +
    CreateDate.  We group by those three fields and keep only the best-status
    entry (Processed/Completed wins over Failed/Initialized).
    Single-entry groups pass through unchanged.
    """
    groups: Dict[str, List[Dict[str, Any]]] = {}

    for order in orders:
        branch = str(order.get("_queried_branch") or order.get("BranchUUID") or "")
        order_number = str(order.get("OrderNumber") or "")
        create_date = str(order.get("CreateDate") or "")
        ts_key = create_date[:19]  # YYYY-MM-DDTHH:MM:SS — drops ms / Z suffix

        if order_number and ts_key:
            key = f"{branch}|{order_number}|{ts_key}"
        else:
            key = str(order.get("ItemId") or id(order))

        groups.setdefault(key, []).append(order)

    result: List[Dict[str, Any]] = []
    deduped_count = 0
    for group in groups.values():
        if len(group) == 1:
            result.append(group[0])
        else:
            deduped_count += len(group) - 1
            best = min(
                group,
                key=lambda o: _STATUS_RANK.get(str(o.get("OrderStatus") or ""), 99),
            )
            result.append(best)

    if deduped_count:
        import logging as _logging
        _logging.getLogger("accounting_export").info(
            "Retry dedup: removed %d duplicate order(s)", deduped_count
        )

    return result


def _summarize_orders(
    orders: Iterable[Dict[str, Any]]
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    buckets: Dict[str, SummaryBucket] = {"All Branches": SummaryBucket()}
    detail_rows: List[Dict[str, Any]] = []
    seen_orders: set[str] = set()

    for order in orders:
        item_id = str(order.get("ItemId") or "")
        # Use the branch UUID we queried for (reliable) rather than whatever
        # OMK echoes back in the response field (which may reflect the refund
        # terminal rather than the original order's branch).
        queried_branch_id = str(order.get("_queried_branch") or order.get("BranchUUID") or "")
        response_branch_id = str(order.get("BranchUUID") or "")
        branch_name = _branch_label(queried_branch_id)
        bucket = buckets.setdefault(branch_name, SummaryBucket())

        if not item_id or item_id in seen_orders:
            continue
        seen_orders.add(item_id)

        discounts_total = _decimal(order.get("DiscountAmount"))
        tips_total = _decimal(order.get("TipAmount"))
        kiosk_sales = _decimal(order.get("SubTotal"))
        gross_total = _decimal(order.get("TotalAmount")) + discounts_total + tips_total
        net_total = kiosk_sales + discounts_total + tips_total
        payment_method = str(order.get("PaymentMethod") or "Unknown")

        for target in (buckets["All Branches"], bucket):
            target.gross_total += gross_total
            target.net_total += net_total
            target.discounts_total += discounts_total
            target.tips_total += tips_total
            target.payment_methods[payment_method] += gross_total

        order_products = order.get("OrderProducts") or []
        if not isinstance(order_products, list):
            order_products = []

        rate_gross_map: Dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
        for product in order_products:
            if not isinstance(product, dict):
                continue

            gross_line, net_line, tax_line, rate = _line_amounts(product)
            rate_bucket = _bucket_for_rate(rate)

            if rate_bucket is not None:
                rate_gross_map[rate_bucket] += gross_line

            refund_gross, refund_net, refund_tax, refund_rate, refunded_qty = _refund_amounts(product)
            refund_bucket = _bucket_for_rate(refund_rate)

            for target in (buckets["All Branches"], bucket):
                target.refunds_gross += refund_gross
                target.refunds_net += refund_net
                if refunded_qty > 0:
                    target.refund_count += 1
                if refund_bucket == "81":
                    target.refund_revenue_81 += refund_gross
                    target.refund_tax_81 += refund_tax
                elif refund_bucket == "26":
                    target.refund_revenue_26 += refund_gross
                    target.refund_tax_26 += refund_tax

            detail_rows.append(
                {
                    "BranchName": branch_name,
                    "BranchUUID": queried_branch_id,
                    "OMK_BranchUUID": response_branch_id,
                    "DeviceType": (order.get("Device") or {}).get("DeviceType"),
                    "ItemId": order.get("ItemId"),
                    "CreateDate": order.get("CreateDate"),
                    "OrderNumber": order.get("OrderNumber"),
                    "OrderType": order.get("OrderType"),
                    "OrderStatus": order.get("OrderStatus"),
                    "PaymentMethod": order.get("PaymentMethod"),
                    "PaymentReferenceId": order.get("PaymentReferenceId"),
                    "TotalAmount": _money(_decimal(order.get("TotalAmount"))),
                    "SubTotal": _money(_decimal(order.get("SubTotal"))),
                    "DiscountAmount": _money(_decimal(order.get("DiscountAmount"))),
                    "TaxAmount": _money(_decimal(order.get("TaxAmount"))),
                    "TipAmount": _money(_decimal(order.get("TipAmount"))),
                    "DeliveryCost": _money(_decimal(order.get("DeliveryCost"))),
                    "ProductName": product.get("Name"),
                    "ProductId": product.get("ProductId"),
                    "ProductVariationName": product.get("ProductVariationName"),
                    "CategoryName": product.get("CategoryName"),
                    "Quantity": float(_decimal(product.get("Quantity"))),
                    "UnitPrice": _money(_decimal(product.get("UnitPrice"))),
                    "LineDiscountAmount": _money(_decimal(product.get("DiscountAmount"))),
                    "TaxRate": float(rate),
                    "LineGrossRevenue": _money(gross_line),
                    "LineNetRevenue": _money(net_line),
                    "LineTaxAmount": _money(tax_line),
                    "ModifierTotal": _money(_modifier_total(product)),
                    "RefundedQuantity": float(refunded_qty),
                    "RefundGrossAmount": _money(refund_gross),
                    "RefundNetAmount": _money(refund_net),
                    "RefundTaxAmount": _money(refund_tax),
                    "RefundedReason": "",
                    "RefundedTime": "",
                }
            )

        order_total_amount = _decimal(order.get("TotalAmount"))
        line_total_amount = sum(rate_gross_map.values(), Decimal("0"))
        if line_total_amount > 0 and order_total_amount >= 0:
            scaling_factor = order_total_amount / line_total_amount
            for rate_bucket, gross_amount in list(rate_gross_map.items()):
                rate_gross_map[rate_bucket] = gross_amount * scaling_factor

        for target in (buckets["All Branches"], bucket):
            gross_81 = rate_gross_map.get("81", Decimal("0"))
            gross_26 = rate_gross_map.get("26", Decimal("0"))

            target.revenue_81 += gross_81
            target.revenue_26 += gross_26

            if gross_81 > 0:
                net_81 = gross_81 / (Decimal("1") + (Decimal("8.1") / Decimal("100")))
                target.tax_due_81 += gross_81 - net_81
            if gross_26 > 0:
                net_26 = gross_26 / (Decimal("1") + (Decimal("2.6") / Decimal("100")))
                target.tax_due_26 += gross_26 - net_26

    summary_rows: List[Dict[str, Any]] = []
    storno_rows: List[Dict[str, Any]] = []
    for scope, bucket in buckets.items():
        summary_rows.append(
            {
                "Scope": scope,
                "Brutto Verkäufe (vor Storno)": _money(bucket.gross_total),
                "Stornierungen / Gutschriften (Brutto)": _money(bucket.refunds_gross),
                "Brutto Total (nach Storno)": _money(bucket.gross_total - bucket.refunds_gross),
                "Netto Total (nach Storno)": _money(bucket.net_total - bucket.refunds_net),
                "Total zu zahlende Mehrwertsteuer 8.1% (nach Storno)": _money(
                    bucket.tax_due_81 - bucket.refund_tax_81
                ),
                "Total zu zahlende Mehrwertsteuer 2.6% (nach Storno)": _money(
                    bucket.tax_due_26 - bucket.refund_tax_26
                ),
                "Total Umsatz mit 8.1 % Mehrwertsteuer (nach Storno)": _money(
                    bucket.revenue_81 - bucket.refund_revenue_81
                ),
                "Total Umsatz mit 2.6 % Mehrwertsteuer (nach Storno)": _money(
                    bucket.revenue_26 - bucket.refund_revenue_26
                ),
                "Total Rabatte": _money(bucket.discounts_total),
                "Total Trinkgelder": _money(bucket.tips_total),
            }
        )
        # Full refund breakdown so the deduction above is transparent — always
        # one row per scope, even when every value is 0.
        storno_rows.append(
            {
                "Scope": scope,
                "Anzahl Stornierungen": bucket.refund_count,
                "Storno Brutto": _money(bucket.refunds_gross),
                "Storno Netto": _money(bucket.refunds_net),
                "Storno Mehrwertsteuer 8.1%": _money(bucket.refund_tax_81),
                "Storno Mehrwertsteuer 2.6%": _money(bucket.refund_tax_26),
                "Storno Umsatz 8.1 % (Brutto)": _money(bucket.refund_revenue_81),
                "Storno Umsatz 2.6 % (Brutto)": _money(bucket.refund_revenue_26),
            }
        )

    # Union of every payment method seen across all branches, so each scope
    # reports every method — explicitly as 0 when it never occurred there.
    all_payment_methods = sorted(
        {method for bucket in buckets.values() for method in bucket.payment_methods}
    )
    payment_rows: List[Dict[str, Any]] = []
    for scope, bucket in buckets.items():
        for payment_method in all_payment_methods:
            total = bucket.payment_methods.get(payment_method, Decimal("0"))
            payment_rows.append(
                {
                    "Scope": scope,
                    "PaymentMethod": payment_method,
                    "Total Zahlungsmethode": _money(total),
                }
            )

    return summary_rows, storno_rows, payment_rows, detail_rows


def _apply_currency_format(sheet, start_row: int, end_row: int, columns: Iterable[int]) -> None:
    for row in range(start_row, end_row + 1):
        for column in columns:
            sheet.cell(row=row, column=column).number_format = '#,##0.00'


def _style_table_header(sheet, row: int, fill_color: str = "D26A1D") -> None:
    fill = PatternFill(fill_type="solid", fgColor=fill_color)
    for cell in sheet[row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _extract_de_name(raw: str) -> str:
    """Extract the German name from a JSON-encoded multilang string, or return raw."""
    import json
    raw = (raw or "").strip()
    if raw.startswith("{"):
        try:
            parsed = json.loads(raw)
            return (parsed.get("de") or parsed.get("en") or raw).strip()
        except Exception:
            pass
    return raw


def build_young_boys_product_workbook(req: RunRequest) -> Tuple[bytes, List[str], List[ProgressEvent]]:
    orders, events, errors = fetch_accounting_orders(req)
    any_ok = any(event.status == "ok" for event in events)
    if not orders and errors and not any_ok:
        raise RuntimeError("Export failed - no rows collected. Errors: " + "; ".join(errors[:5]))

    # Aggregate quantities: branch_name -> product_name -> total_qty
    branch_products: Dict[str, DefaultDict[str, Decimal]] = {}
    seen_orders: set[str] = set()

    for order in orders:
        item_id = str(order.get("ItemId") or "")
        if not item_id or item_id in seen_orders:
            continue
        seen_orders.add(item_id)

        queried_branch_id = str(order.get("_queried_branch") or order.get("BranchUUID") or "")
        branch_name = _branch_label(queried_branch_id)

        if branch_name not in branch_products:
            branch_products[branch_name] = defaultdict(lambda: Decimal("0"))

        order_products = order.get("OrderProducts") or []
        if not isinstance(order_products, list):
            continue

        for product in order_products:
            if not isinstance(product, dict):
                continue
            name = _extract_de_name(product.get("Name") or "")
            qty = _decimal(product.get("Quantity"))
            if name and qty > 0:
                branch_products[branch_name][name] += qty

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Produkte nach Box"

    current_row = 1
    for branch_name in BSC_YOUNG_BOYS_BRANCHES.values():
        # Branch header
        sheet.cell(row=current_row, column=1, value=branch_name)
        sheet.cell(row=current_row, column=1).font = Font(bold=True, size=13, color="FFFFFF")
        sheet.cell(row=current_row, column=1).fill = PatternFill(fill_type="solid", fgColor="1A3A6B")
        sheet.cell(row=current_row, column=2).fill = PatternFill(fill_type="solid", fgColor="1A3A6B")
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        current_row += 1

        # Column headers
        sheet.cell(row=current_row, column=1, value="Produkt")
        sheet.cell(row=current_row, column=2, value="Menge")
        _style_table_header(sheet, current_row, fill_color="2E5BA8")
        current_row += 1

        products = branch_products.get(branch_name, {})
        for product_name, qty in sorted(products.items(), key=lambda x: -x[1]):
            sheet.cell(row=current_row, column=1, value=product_name)
            sheet.cell(row=current_row, column=2, value=float(qty))
            current_row += 1

        current_row += 1  # blank row between branches

    sheet.column_dimensions["A"].width = 40
    sheet.column_dimensions["B"].width = 10

    output = BytesIO()
    workbook.save(output)
    return output.getvalue(), errors, events


def build_young_boys_tcpos_export(req: RunRequest) -> Tuple[bytes, List[str], List[ProgressEvent]]:
    from .tcpos_mapping import lookup_tcpos

    # BOX -> Profit Center
    BOX_TO_PC = {
        "BOX 6 - Pizza":  1006,
        "BOX 7 - Pommes": 1007,
        "BOX 8 - Grill":  1008,
        "BOX 9 - Döner":  1009,
    }
    MWST = 2

    orders, events, errors = fetch_accounting_orders(req)
    any_ok = any(event.status == "ok" for event in events)
    if not orders and errors and not any_ok:
        raise RuntimeError("Export failed - no rows collected. Errors: " + "; ".join(errors[:5]))

    # Aggregate: branch -> artikel_nr (or fallback omk_name) -> {qty, revenue, tcpos_bez}
    # Keying by artikel_nr ensures multiple OMK variants of the same TCPOS article are merged.
    branch_products: Dict[str, Dict[str, Dict]] = {
        b: {} for b in BSC_YOUNG_BOYS_BRANCHES.values()
    }
    unmapped: set[str] = set()
    seen_orders: set[str] = set()

    for order in orders:
        item_id = str(order.get("ItemId") or "")
        if not item_id or item_id in seen_orders:
            continue
        seen_orders.add(item_id)

        queried_branch_id = str(order.get("_queried_branch") or order.get("BranchUUID") or "")
        branch_name = _branch_label(queried_branch_id)
        if branch_name not in branch_products:
            continue

        for product in (order.get("OrderProducts") or []):
            if not isinstance(product, dict):
                continue
            name_de = _extract_de_name(product.get("Name") or "")
            qty = _decimal(product.get("Quantity"))
            if not name_de or qty <= 0:
                continue

            nr, tcpos_bez = lookup_tcpos(name_de)
            if nr is None:
                unmapped.add(name_de)

            # Aggregate by artikel_nr when available, else fall back to omk name
            agg_key = nr if nr else f"__unmapped__{name_de}"
            agg = branch_products[branch_name].setdefault(
                agg_key,
                {"nr": nr or "", "bez": tcpos_bez or name_de, "qty": Decimal("0"), "revenue": Decimal("0")},
            )
            agg["qty"] += qty
            agg["revenue"] += _decimal(product.get("UnitPrice")) * qty

    if unmapped:
        errors.append("Kein TCPOS-Mapping gefunden für: " + ", ".join(sorted(unmapped)))

    workbook = Workbook()
    ws = workbook.active
    ws.title = "TCPOS Export"

    # Header row
    cols = ["Artikel Nr.", "Artikel Bezeichnung", "Total Umsatz", "Stück", "MWST", "Profit Center"]
    ws.append(cols)
    _style_table_header(ws, 1, fill_color="1A3A6B")
    ws.row_dimensions[1].height = 18

    for branch_name in BSC_YOUNG_BOYS_BRANCHES.values():
        profit_center = BOX_TO_PC.get(branch_name, 0)
        products = branch_products.get(branch_name, {})
        for agg_key, agg in sorted(products.items(), key=lambda x: x[1]["bez"]):
            ws.append([
                agg["nr"],
                agg["bez"],
                _money(agg["revenue"]),
                int(agg["qty"]),
                MWST,
                profit_center,
            ])

    # Format
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 14
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    if ws.max_row > 1:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = '#,##0.00'

    output = BytesIO()
    workbook.save(output)
    return output.getvalue(), errors, events


_YB_UPSELL_KEYWORDS = ("mandelbärli", "ybächer")


def build_young_boys_match_report(req: RunRequest) -> Tuple[bytes, List[str], List[ProgressEvent]]:
    """Generate BSC Young Boys Match Report Excel — 1:1 match of ORDERMONKEY_Match_Report PDF."""
    orders, events, errors = fetch_accounting_orders(req)
    any_ok = any(event.status == "ok" for event in events)
    if not orders and errors and not any_ok:
        raise RuntimeError("Export failed - no rows collected. Errors: " + "; ".join(errors[:5]))

    orders = _deduplicate_retry_orders(orders)
    branches = list(BSC_YOUNG_BOYS_BRANCHES.values())

    # ── Data aggregation ──────────────────────────────────────────────────────
    branch_stats: Dict[str, Dict[str, Decimal]] = {
        b: {"kiosk_count": Decimal("0"), "kiosk_revenue": Decimal("0"),
            "app_count": Decimal("0"), "app_revenue": Decimal("0")}
        for b in branches
    }
    upsell_data: Dict[str, Dict[str, Dict[str, Decimal]]] = {}
    upsell_total_revenue = Decimal("0")
    seen_orders: set = set()

    for order in orders:
        item_id = str(order.get("ItemId") or "")
        if not item_id or item_id in seen_orders:
            continue
        seen_orders.add(item_id)

        queried_branch_id = str(order.get("_queried_branch") or order.get("BranchUUID") or "")
        branch_name = _branch_label(queried_branch_id)
        if branch_name not in branch_stats:
            continue

        device_raw = str((order.get("Device") or {}).get("DeviceType") or "")
        is_kiosk = device_raw.upper() == "KIOSK"
        revenue = _decimal(order.get("TotalAmount"))

        if is_kiosk:
            branch_stats[branch_name]["kiosk_count"] += Decimal("1")
            branch_stats[branch_name]["kiosk_revenue"] += revenue
        else:
            branch_stats[branch_name]["app_count"] += Decimal("1")
            branch_stats[branch_name]["app_revenue"] += revenue

        for product in (order.get("OrderProducts") or []):
            if not isinstance(product, dict):
                continue
            prod_name_raw = str(product.get("Name") or "")
            name_de = _extract_de_name(prod_name_raw)  # parse multilang JSON first
            if any(kw in name_de.lower() for kw in _YB_UPSELL_KEYWORDS):
                qty = _decimal(product.get("Quantity") or "1")
                unit_price = _decimal(product.get("UnitPrice") or "0")
                line_rev = qty * unit_price
                upsell_total_revenue += line_rev
                pd = upsell_data.setdefault(name_de, {})
                bd = pd.setdefault(branch_name, {"count": Decimal("0"), "revenue": Decimal("0")})
                bd["count"] += qty
                bd["revenue"] += line_rev

    # ── Totals ────────────────────────────────────────────────────────────────
    _D = Decimal
    _z = _D("0")

    total_kiosk_count = sum(branch_stats[b]["kiosk_count"] for b in branches)
    total_kiosk_rev   = sum(branch_stats[b]["kiosk_revenue"] for b in branches)
    total_app_count   = sum(branch_stats[b]["app_count"] for b in branches)
    total_app_rev     = sum(branch_stats[b]["app_revenue"] for b in branches)
    total_count = total_kiosk_count + total_app_count
    total_rev   = total_kiosk_rev + total_app_rev

    avg_order  = _money(total_rev / total_count) if total_count > _z else 0.0
    avg_kiosk  = _money(total_kiosk_rev / total_kiosk_count) if total_kiosk_count > _z else 0.0
    avg_app    = _money(total_app_rev / total_app_count) if total_app_count > _z else 0.0
    app_aufschlag = round(avg_app - avg_kiosk, 2)
    avg_per_box   = _money(total_rev / _D(len(branches))) if branches else 0.0
    avg_orders_per_box = round(float(total_count) / len(branches), 1) if branches else 0.0

    top_box = (max(branches,
                   key=lambda b: branch_stats[b]["kiosk_revenue"] + branch_stats[b]["app_revenue"])
               if branches else "–")
    top_box_rev = (_money(branch_stats[top_box]["kiosk_revenue"] + branch_stats[top_box]["app_revenue"])
                   if top_box != "–" else 0.0)

    total_upsell_pieces = sum(bd["count"] for pd in upsell_data.values() for bd in pd.values())
    avg_upsell_per_piece = (_money(upsell_total_revenue / total_upsell_pieces)
                            if total_upsell_pieces > _z else 0.0)

    # ── Date strings (fix: strip time component before parsing) ───────────────
    spieltag_str = ""
    if req.startDate:
        try:
            date_part = req.startDate.split("T")[0]
            spieltag_str = datetime.strptime(date_part, "%Y-%m-%d").strftime("%d.%m.%Y")
        except Exception:
            spieltag_str = req.startDate
    gen_str = datetime.now().strftime("%d.%m.%Y %H:%M")

    def _pct(part: Decimal, whole: Decimal) -> float:
        return round(float(part / whole * 100), 1) if whole > _z else 0.0

    def _chf_str(v: float) -> str:
        return f"CHF {v:,.2f}".replace(",", "'")

    # ── Colors (matching PDF exactly) ─────────────────────────────────────────
    RED       = "CC0000"  # header / section bars
    WHITE     = "FFFFFF"
    DARK      = "1A1A1A"  # total rows bg / near-black text
    PINK      = "FFE8E8"  # column header rows
    LIGHT_KPI = "F2F2F2"  # KPI tiles 2-6
    ALT_A     = "F5F5F5"  # alternating row A
    ALT_B     = "FFFFFF"  # alternating row B
    GRAY_TEXT = "888888"

    # ── Workbook / sheet setup ─────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Match Report"
    ws.sheet_view.showGridLines = False

    # A4 portrait, fit to 1 page
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Columns A-J (5 left + 5 right)
    col_widths = [23, 9, 14, 8, 9,   23, 9, 14, 8, 9]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _fill(r, c1, c2, color):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = PatternFill(fill_type="solid", fgColor=color)

    def _mc(r, c1, c2, value=None):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        return ws.cell(row=r, column=c1, value=value)

    def _rh(r, h):
        ws.row_dimensions[r].height = h

    def _sec_hdr(r, c1, c2, title):
        _fill(r, c1, c2, RED)
        cell = _mc(r, c1, c2, title)
        cell.font = Font(bold=True, size=10, color=WHITE)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    def _col_hdr(r, col, lbl, first=False):
        cell = ws.cell(row=r, column=col, value=lbl)
        cell.font = Font(bold=True, size=8, color=DARK)
        cell.fill = PatternFill(fill_type="solid", fgColor=PINK)
        cell.alignment = Alignment(
            horizontal="left" if first else "right",
            vertical="center",
            indent=1 if first else 0,
        )

    def _data_cell(r, col, val, bold=False, color=DARK, fmt=None, indent=False, right=False):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font = Font(bold=bold, size=9, color=color)
        if indent:
            cell.alignment = Alignment(indent=1)
        elif right:
            cell.alignment = Alignment(horizontal="right")
        if fmt:
            cell.number_format = fmt

    # ── Row 1-2: Header ───────────────────────────────────────────────────────
    _fill(row, 1, 10, RED)
    _rh(row, 30)
    cell = _mc(row, 1, 10, "BSC YOUNG BOYS · MATCH REPORT")
    cell.font = Font(bold=True, size=16, color=WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    _fill(row, 1, 10, DARK)
    _rh(row, 15)
    cell = _mc(row, 1, 10,
               f"Spieltag {spieltag_str}   ·   Kiosk & Mobile App   ·   Erstellt {gen_str}   ·   ORDERMONKEY")
    cell.font = Font(size=8, color="AAAAAA")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    # ── KPI tiles row 1 (3 sub-rows: label / value / subtitle) ───────────────
    # Tile 1: RED bg (TOTAL UMSATZ). Tiles 2-3: light bg.
    tile_spans = [(1, 3), (4, 7), (8, 10)]
    kpi1 = [
        {
            "label": "TOTAL UMSATZ (CHF)",
            "value": _chf_str(_money(total_rev)),
            "sub":   "Kiosk & Mobile App · inkl. MwSt.",
            "bg": RED, "fg": WHITE, "label_c": "FFCCCC", "sub_c": "FFCCCC",
        },
        {
            "label": "BESTELLUNGEN TOTAL",
            "value": str(int(total_count)),
            "sub":   f"{int(total_kiosk_count)} Kiosk · {int(total_app_count)} App",
            "bg": LIGHT_KPI, "fg": DARK, "label_c": GRAY_TEXT, "sub_c": GRAY_TEXT,
        },
        {
            "label": "Ø BESTELLWERT",
            "value": _chf_str(avg_order),
            "sub":   f"Kiosk {_chf_str(avg_kiosk)} · App {_chf_str(avg_app)}",
            "bg": LIGHT_KPI, "fg": DARK, "label_c": GRAY_TEXT, "sub_c": GRAY_TEXT,
        },
    ]

    _rh(row, 12)
    for (c1, c2), t in zip(tile_spans, kpi1):
        _fill(row, c1, c2, t["bg"])
        cell = _mc(row, c1, c2, t["label"])
        cell.font = Font(bold=True, size=7, color=t["label_c"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    _rh(row, 24)
    for (c1, c2), t in zip(tile_spans, kpi1):
        _fill(row, c1, c2, t["bg"])
        cell = _mc(row, c1, c2, t["value"])
        cell.font = Font(bold=True, size=14, color=t["fg"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    _rh(row, 12)
    for (c1, c2), t in zip(tile_spans, kpi1):
        _fill(row, c1, c2, t["bg"])
        cell = _mc(row, c1, c2, t["sub"])
        cell.font = Font(size=7, color=t["sub_c"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    # thin gap between tile rows
    _fill(row, 1, 10, WHITE)
    _rh(row, 5)
    row += 1

    # ── KPI tiles row 2 ───────────────────────────────────────────────────────
    kpi2 = [
        {
            "label": "TOP BOX NACH UMSATZ",
            "value": top_box,
            "sub":   _chf_str(top_box_rev),
            "bg": LIGHT_KPI, "fg": DARK, "label_c": GRAY_TEXT, "sub_c": GRAY_TEXT,
        },
        {
            "label": "Ø UMSATZ PRO BOX (CHF)",
            "value": _chf_str(avg_per_box),
            "sub":   f"über {len(branches)} Boxen",
            "bg": LIGHT_KPI, "fg": DARK, "label_c": GRAY_TEXT, "sub_c": GRAY_TEXT,
        },
        {
            "label": "UPSELL UMSATZ (CHF)",
            "value": _chf_str(_money(upsell_total_revenue)),
            "sub":   f"{int(total_upsell_pieces)} Stück · Ø {_chf_str(avg_upsell_per_piece)}",
            "bg": LIGHT_KPI, "fg": DARK, "label_c": GRAY_TEXT, "sub_c": GRAY_TEXT,
        },
    ]

    _rh(row, 12)
    for (c1, c2), t in zip(tile_spans, kpi2):
        _fill(row, c1, c2, t["bg"])
        cell = _mc(row, c1, c2, t["label"])
        cell.font = Font(bold=True, size=7, color=t["label_c"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    _rh(row, 24)
    for (c1, c2), t in zip(tile_spans, kpi2):
        _fill(row, c1, c2, t["bg"])
        cell = _mc(row, c1, c2, t["value"])
        cell.font = Font(bold=True, size=14, color=t["fg"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    _rh(row, 12)
    for (c1, c2), t in zip(tile_spans, kpi2):
        _fill(row, c1, c2, t["bg"])
        cell = _mc(row, c1, c2, t["sub"])
        cell.font = Font(size=7, color=t["sub_c"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    _rh(row, 8)
    row += 1

    # ── UMSATZ PRO BOX (cols 1-5) + KANAL-SPLIT (cols 6-10) ─────────────────
    _sec_hdr(row, 1, 5, "UMSATZ PRO BOX")
    _sec_hdr(row, 6, 10, "KANAL-SPLIT")
    _rh(row, 20)
    row += 1

    for col, lbl in zip([1, 2, 3, 4, 5], ["Box", "Anzahl", "Warenwert CHF", "Anteil %", "Ø Bestellwert"]):
        _col_hdr(row, col, lbl, first=(col == 1))
    for col, lbl in zip([6, 7, 8, 9, 10], ["Kanal", "Anzahl", "Warenwert CHF", "Anteil %", "Ø Bestellwert"]):
        _col_hdr(row, col, lbl, first=(col == 6))
    _rh(row, 14)
    row += 1

    kanal_rows = [
        ("Kiosk",      int(total_kiosk_count), _money(total_kiosk_rev),
         _pct(total_kiosk_rev, total_rev), avg_kiosk),
        ("Mobile App", int(total_app_count),   _money(total_app_rev),
         _pct(total_app_rev, total_rev),   avg_app),
    ]

    n_section = max(len(branches), len(kanal_rows))
    for i in range(n_section):
        bg = ALT_A if i % 2 == 0 else ALT_B
        _rh(row, 14)

        if i < len(branches):
            branch = branches[i]
            b = branch_stats[branch]
            bc = b["kiosk_count"] + b["app_count"]
            br = b["kiosk_revenue"] + b["app_revenue"]
            avg_b = round(_money(br) / float(bc), 2) if bc > _z else 0.0
            _fill(row, 1, 5, bg)
            _data_cell(row, 1, branch, color=DARK, indent=True)
            _data_cell(row, 2, int(bc), color=DARK, right=True)
            _data_cell(row, 3, _money(br), color=DARK, right=True, fmt='#,##0.00')
            _data_cell(row, 4, _pct(br, total_rev), color=DARK, right=True, fmt='0.0"%"')
            _data_cell(row, 5, avg_b, color=DARK, right=True, fmt='#,##0.00')
        else:
            _fill(row, 1, 5, ALT_B)

        if i < len(kanal_rows):
            k = kanal_rows[i]
            _fill(row, 6, 10, bg)
            _data_cell(row, 6, k[0], color=DARK, indent=True)
            _data_cell(row, 7, k[1], color=DARK, right=True)
            _data_cell(row, 8, k[2], color=DARK, right=True, fmt='#,##0.00')
            _data_cell(row, 9, k[3], color=DARK, right=True, fmt='0.0"%"')
            _data_cell(row, 10, k[4], color=DARK, right=True, fmt='#,##0.00')
        else:
            _fill(row, 6, 10, ALT_B)

        row += 1

    # Total row — DARK background, WHITE text
    total_avg_order = round(float(_money(total_rev)) / float(total_count), 2) if total_count > _z else 0.0
    _fill(row, 1, 5, DARK)
    _data_cell(row, 1, "Total", bold=True, color=WHITE, indent=True)
    _data_cell(row, 2, int(total_count), bold=True, color=WHITE, right=True)
    _data_cell(row, 3, _money(total_rev), bold=True, color=WHITE, right=True, fmt='#,##0.00')
    _data_cell(row, 4, 100.0, bold=True, color=WHITE, right=True, fmt='0.0"%"')
    _data_cell(row, 5, total_avg_order, bold=True, color=WHITE, right=True, fmt='#,##0.00')

    _fill(row, 6, 10, DARK)
    _data_cell(row, 6, "Total", bold=True, color=WHITE, indent=True)
    _data_cell(row, 7, int(total_count), bold=True, color=WHITE, right=True)
    _data_cell(row, 8, _money(total_rev), bold=True, color=WHITE, right=True, fmt='#,##0.00')
    _data_cell(row, 9, 100.0, bold=True, color=WHITE, right=True, fmt='0.0"%"')
    _data_cell(row, 10, total_avg_order, bold=True, color=WHITE, right=True, fmt='#,##0.00')
    _rh(row, 15)
    row += 1

    _rh(row, 8)
    row += 1

    # ── UPSELL PRODUKTE (cols 1-5) + WEITERE KENNZAHLEN (cols 6-10) ──────────
    _sec_hdr(row, 1, 5, "UPSELL PRODUKTE")
    _sec_hdr(row, 6, 10, "WEITERE KENNZAHLEN")
    _rh(row, 20)
    row += 1

    # UPSELL: product × box matrix — cols: Produkt | Box6 | Box7 | Box8 | Box9
    branch_short = [b.split(" ", 1)[-1] if " " in b else b for b in branches]  # "Pizza", "Pommes"…
    upsell_hdr_labels = ["Produkt"] + branch_short
    for col, lbl in zip([1, 2, 3, 4, 5], upsell_hdr_labels[:5]):
        _col_hdr(row, col, lbl, first=(col == 1))

    # WEITERE: Kennzahl | Wert (merge cols 6-7 / 8-10)
    for col in [6, 7, 8, 9, 10]:
        ws.cell(row=row, column=col).fill = PatternFill(fill_type="solid", fgColor=PINK)
    cell_k = _mc(row, 6, 7, "Kennzahl")
    cell_k.font = Font(bold=True, size=8, color=DARK)
    cell_k.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell_v = _mc(row, 8, 10, "Wert")
    cell_v.font = Font(bold=True, size=8, color=DARK)
    cell_v.alignment = Alignment(horizontal="right", vertical="center")
    _rh(row, 14)
    row += 1

    upsell_products = list(upsell_data.keys())
    box_upsell_qty = {b: Decimal("0") for b in branches}
    for prod_name, pd in upsell_data.items():
        for bn, bd in pd.items():
            if bn in box_upsell_qty:
                box_upsell_qty[bn] += bd["count"]

    weitere_normal = [
        ("Ø Bestellwert",          _chf_str(avg_order)),
        ("Ø Upsell-Wert/Stück",    _chf_str(avg_upsell_per_piece)),
        ("Ø Bestellungen/Box",     f"{avg_orders_per_box:.1f}"),
        ("App-Aufschlag vs. Kiosk", f"CHF {app_aufschlag:+,.2f}".replace(",", "'")),
    ]

    n_data = max(len(upsell_products) if upsell_products else 1, len(weitere_normal))
    for i in range(n_data):
        bg = ALT_A if i % 2 == 0 else ALT_B
        _rh(row, 14)

        # left: upsell matrix
        if not upsell_products and i == 0:
            _fill(row, 1, 5, bg)
            cell = _mc(row, 1, 5, "– keine Upsells –")
            cell.font = Font(size=9, color=GRAY_TEXT, italic=True)
            cell.alignment = Alignment(horizontal="center")
        elif i < len(upsell_products):
            prod = upsell_products[i]
            _fill(row, 1, 5, bg)
            _data_cell(row, 1, prod, color=DARK, indent=True)
            for j, branch in enumerate(branches, start=2):
                qty_val = upsell_data[prod].get(branch, {}).get("count", Decimal("0"))
                val = int(qty_val) if qty_val > _z else "–"
                _data_cell(row, j, val, color=DARK, right=True)
        else:
            _fill(row, 1, 5, ALT_B)

        # right: weitere kennzahlen
        if i < len(weitere_normal):
            _fill(row, 6, 10, bg)
            for col in [6, 7, 8, 9, 10]:
                ws.cell(row=row, column=col).fill = PatternFill(fill_type="solid", fgColor=bg)
            cell_lbl = _mc(row, 6, 7, weitere_normal[i][0])
            cell_lbl.font = Font(size=9, color=DARK)
            cell_lbl.alignment = Alignment(indent=1)
            cell_val = _mc(row, 8, 10, weitere_normal[i][1])
            cell_val.font = Font(bold=True, size=9, color=DARK)
            cell_val.alignment = Alignment(horizontal="right")
        else:
            _fill(row, 6, 10, ALT_B)

        row += 1

    # Bottom totals row: TOTAL UPSELLS (left) + Ø UMSATZ PRO BOX highlighted (right)
    _fill(row, 1, 5, DARK)
    if upsell_products:
        _data_cell(row, 1, "TOTAL UPSELLS", bold=True, color=WHITE, indent=True)
        for j, branch in enumerate(branches, start=2):
            qty = int(box_upsell_qty[branch])
            _data_cell(row, j, qty if qty > 0 else "–", bold=True, color=WHITE, right=True)
    else:
        cell = _mc(row, 1, 5, "")
        _fill(row, 1, 5, DARK)

    _fill(row, 6, 10, DARK)
    for col in [6, 7, 8, 9, 10]:
        ws.cell(row=row, column=col).fill = PatternFill(fill_type="solid", fgColor=DARK)
    cell_lbl = _mc(row, 6, 7, "Ø UMSATZ PRO BOX")
    cell_lbl.font = Font(bold=True, size=9, color=WHITE)
    cell_lbl.alignment = Alignment(indent=1)
    cell_val = _mc(row, 8, 10, _chf_str(avg_per_box))
    cell_val.font = Font(bold=True, size=9, color=WHITE)
    cell_val.alignment = Alignment(horizontal="right")
    _rh(row, 15)
    row += 1

    _rh(row, 8)
    row += 1

    # ── DETAIL PRO BOX · KIOSK vs. MOBILE APP (full width) ───────────────────
    _fill(row, 1, 10, RED)
    cell = _mc(row, 1, 10, "DETAIL PRO BOX · KIOSK vs. MOBILE APP")
    cell.font = Font(bold=True, size=10, color=WHITE)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    _rh(row, 20)
    row += 1

    detail_hdrs = ["Box", "Kiosk Anz.", "Kiosk CHF", "App Anz.", "App CHF",
                   "Total Anz.", "Total CHF", "Ø Bestellwert", "% Umsatz", "App-Anteil"]
    for col, lbl in enumerate(detail_hdrs, start=1):
        _col_hdr(row, col, lbl, first=(col == 1))
    _rh(row, 14)
    row += 1

    for i, branch in enumerate(branches):
        bg = ALT_A if i % 2 == 0 else ALT_B
        b = branch_stats[branch]
        kc = int(b["kiosk_count"]); kr = _money(b["kiosk_revenue"])
        ac = int(b["app_count"]);   ar = _money(b["app_revenue"])
        tc = kc + ac; tr = round(kr + ar, 2)
        avg_o   = round(tr / tc, 2) if tc > 0 else 0.0
        pct_u   = _pct(b["kiosk_revenue"] + b["app_revenue"], total_rev)
        app_ant = _pct(b["app_revenue"], b["kiosk_revenue"] + b["app_revenue"])
        _fill(row, 1, 10, bg)
        for col, val in enumerate([branch, kc, kr, ac, ar, tc, tr, avg_o, pct_u, app_ant], start=1):
            fmt = None
            if col in (3, 5, 7, 8): fmt = '#,##0.00'
            elif col in (9, 10):    fmt = '0.0"%"'
            _data_cell(row, col, val, color=DARK, fmt=fmt,
                       indent=(col == 1), right=(col != 1))
        _rh(row, 14)
        row += 1

    # Total row — DARK background, WHITE text
    _fill(row, 1, 10, DARK)
    t_avg    = round(float(_money(total_rev)) / float(total_count), 2) if total_count > _z else 0.0
    t_app_ant = _pct(total_app_rev, total_rev)
    totals = ["Total", int(total_kiosk_count), _money(total_kiosk_rev),
              int(total_app_count), _money(total_app_rev),
              int(total_count), _money(total_rev), t_avg, 100.0, t_app_ant]
    for col, val in enumerate(totals, start=1):
        fmt = None
        if col in (3, 5, 7, 8): fmt = '#,##0.00'
        elif col in (9, 10):    fmt = '0.0"%"'
        _data_cell(row, col, val, bold=True, color=WHITE, fmt=fmt,
                   indent=(col == 1), right=(col != 1))
    _rh(row, 15)
    row += 1

    # ── Footer ────────────────────────────────────────────────────────────────
    _fill(row, 1, 10, RED)
    _rh(row, 3)
    row += 1
    _rh(row, 13)
    cell = _mc(row, 1, 10,
               "erstellt durch ORDERMONKEY · «Anzahl» = Bestellungen, alle Werte inkl. MwSt.")
    cell.font = Font(size=7, color=GRAY_TEXT, italic=True)
    cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A4"

    output = BytesIO()
    wb.save(output)
    return output.getvalue(), errors, events


# ─────────────────────────────────────────────────────────────────────────────
#  BSC Young Boys – Season Dashboard  (JSON, no file)
# ─────────────────────────────────────────────────────────────────────────────

def build_young_boys_season_dashboard(req: RunRequest) -> Dict[str, Any]:
    """Return per-match KPI data as a plain dict for the frontend dashboard."""
    from datetime import timezone
    try:
        import zoneinfo
        _TZ = zoneinfo.ZoneInfo("Europe/Zurich")
    except Exception:
        _TZ = None  # fallback: treat as UTC

    WEEKDAYS_DE = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag", "Sonntag"]
    BOX_ORDER   = list(BSC_YOUNG_BOYS_BRANCHES.values())  # canonical box order

    orders, events, errors = fetch_accounting_orders(req)
    any_ok = any(e.status == "ok" for e in events)
    if not orders and errors and not any_ok:
        return {"matches": [], "season": {}, "errors": errors}

    orders = _deduplicate_retry_orders(orders)
    seen_orders: set = set()

    # date_str -> match bucket
    matches: Dict[str, Dict] = {}

    for order in orders:
        item_id = str(order.get("ItemId") or "")
        if not item_id or item_id in seen_orders:
            continue
        seen_orders.add(item_id)

        queried_branch_id = str(order.get("_queried_branch") or order.get("BranchUUID") or "")
        branch_name = _branch_label(queried_branch_id)
        if branch_name not in BSC_YOUNG_BOYS_BRANCHES.values():
            continue

        # Parse timestamp → Swiss local date
        create_raw = str(order.get("CreateDate") or "")
        try:
            dt_utc = datetime.fromisoformat(create_raw.replace("Z", "+00:00"))
            dt_local = dt_utc.astimezone(_TZ) if _TZ else dt_utc
        except Exception:
            continue

        date_str = dt_local.strftime("%Y-%m-%d")
        hour     = dt_local.hour

        revenue  = _decimal(order.get("TotalAmount") or "0")
        device   = str((order.get("Device") or {}).get("DeviceType") or "")
        is_kiosk = device.upper() == "KIOSK"

        m = matches.setdefault(date_str, {
            "date": date_str,
            "weekday": WEEKDAYS_DE[dt_local.weekday()],
            "orders": 0,
            "revenue": Decimal("0"),
            "kiosk_orders": 0, "kiosk_revenue": Decimal("0"),
            "app_orders":   0, "app_revenue":   Decimal("0"),
            "by_hour": defaultdict(int),
            "by_box": {b: {"orders": 0, "revenue": Decimal("0")} for b in BOX_ORDER},
            "products": defaultdict(Decimal),
        })

        m["orders"] += 1
        m["revenue"] += revenue
        if is_kiosk:
            m["kiosk_orders"] += 1
            m["kiosk_revenue"] += revenue
        else:
            m["app_orders"] += 1
            m["app_revenue"] += revenue
        m["by_hour"][hour] += 1
        if branch_name in m["by_box"]:
            m["by_box"][branch_name]["orders"]  += 1
            m["by_box"][branch_name]["revenue"] += revenue

        for product in (order.get("OrderProducts") or []):
            if not isinstance(product, dict):
                continue
            name = _extract_de_name(product.get("Name") or "")
            qty  = _decimal(product.get("Quantity") or "1")
            if name and qty > 0:
                m["products"][name] += qty

    # ── Serialise matches ──────────────────────────────────────────────────────
    result_matches = []
    for date_str, m in sorted(matches.items()):
        n = m["orders"]
        rev = float(m["revenue"])
        k_n = m["kiosk_orders"]
        a_n = m["app_orders"]
        k_rev = float(m["kiosk_revenue"])
        a_rev = float(m["app_revenue"])

        # peak hour (busiest single hour)
        peak_hour = max(m["by_hour"], key=m["by_hour"].get) if m["by_hour"] else None
        peak_orders = m["by_hour"][peak_hour] if peak_hour is not None else 0

        # hourly timeline: list[{hour, orders}] sorted
        timeline = [
            {"hour": h, "orders": cnt}
            for h, cnt in sorted(m["by_hour"].items())
        ]

        top_products = [
            {"name": name, "qty": float(qty)}
            for name, qty in sorted(m["products"].items(), key=lambda x: -x[1])[:5]
        ]

        by_box = {
            box: {"orders": v["orders"], "revenue": round(float(v["revenue"]), 2)}
            for box, v in m["by_box"].items()
        }

        result_matches.append({
            "date": date_str,
            "weekday": m["weekday"],
            "orders": n,
            "revenue": round(rev, 2),
            "avg_order": round(rev / n, 2) if n else 0,
            "kiosk_orders": k_n,
            "app_orders": a_n,
            "kiosk_pct": round(k_n / n * 100, 1) if n else 0,
            "app_pct": round(a_n / n * 100, 1) if n else 0,
            "kiosk_avg": round(k_rev / k_n, 2) if k_n else 0,
            "app_avg": round(a_rev / a_n, 2) if a_n else 0,
            "peak_hour": peak_hour,
            "peak_hour_orders": peak_orders,
            "timeline": timeline,
            "top_products": top_products,
            "by_box": by_box,
        })

    # ── Season summary ─────────────────────────────────────────────────────────
    season: Dict[str, Any] = {}
    if result_matches:
        total_rev    = sum(m["revenue"] for m in result_matches)
        total_orders = sum(m["orders"] for m in result_matches)
        best  = max(result_matches, key=lambda m: m["revenue"])
        worst = min(result_matches, key=lambda m: m["revenue"])

        # aggregate top products across all matches
        all_products: Dict[str, float] = defaultdict(float)
        for m in matches.values():
            for name, qty in m["products"].items():
                all_products[name] += float(qty)
        top_overall = sorted(all_products.items(), key=lambda x: -x[1])[:5]

        total_kiosk = sum(m["kiosk_orders"] for m in result_matches)
        total_app   = sum(m["app_orders"]   for m in result_matches)

        season = {
            "match_count": len(result_matches),
            "total_revenue": round(total_rev, 2),
            "total_orders": total_orders,
            "avg_revenue_per_match": round(total_rev / len(result_matches), 2),
            "avg_order_value": round(total_rev / total_orders, 2) if total_orders else 0,
            "best_match":  {"date": best["date"],  "revenue": best["revenue"],  "weekday": best["weekday"]},
            "worst_match": {"date": worst["date"], "revenue": worst["revenue"], "weekday": worst["weekday"]},
            "kiosk_pct": round(total_kiosk / total_orders * 100, 1) if total_orders else 0,
            "app_pct":   round(total_app   / total_orders * 100, 1) if total_orders else 0,
            "top_products_overall": [{"name": n, "qty": round(q, 1)} for n, q in top_overall],
        }

    return {"matches": result_matches, "season": season, "errors": errors}



def build_accounting_workbook(req: RunRequest) -> Tuple[bytes, List[str], List[ProgressEvent]]:
    orders, events, errors = fetch_accounting_orders(req)
    any_ok = any(event.status == "ok" for event in events)
    if not orders and errors and not any_ok:
        raise RuntimeError("Export failed - no rows collected. Errors: " + "; ".join(errors[:5]))

    orders = _deduplicate_retry_orders(orders)
    summary_rows, storno_rows, payment_rows, detail_rows = _summarize_orders(orders)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.merge_cells("A1:K1")
    summary_sheet["A1"] = f"{_org_label(req.orgId)} Accounting Summary"
    summary_sheet["A1"].font = Font(bold=True, size=16)
    summary_sheet["A1"].alignment = Alignment(horizontal="center")
    summary_sheet["A2"] = "Zeitraum"
    summary_sheet["B2"] = f"{req.startDate or '-'} bis {req.endDate or '-'}"
    summary_sheet["A3"] = "Erstellt am"
    summary_sheet["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary_sheet["A5"] = "Gesamt und pro Betrieb"
    summary_sheet["A5"].font = Font(bold=True, size=12)
    summary_sheet.freeze_panes = "A6"

    summary_columns = [
        "Scope",
        "Brutto Verkäufe (vor Storno)",
        "Stornierungen / Gutschriften (Brutto)",
        "Brutto Total (nach Storno)",
        "Netto Total (nach Storno)",
        "Total zu zahlende Mehrwertsteuer 8.1% (nach Storno)",
        "Total zu zahlende Mehrwertsteuer 2.6% (nach Storno)",
        "Total Umsatz mit 8.1 % Mehrwertsteuer (nach Storno)",
        "Total Umsatz mit 2.6 % Mehrwertsteuer (nach Storno)",
        "Total Rabatte",
        "Total Trinkgelder",
    ]
    summary_table_header_row = 6
    summary_sheet.append(summary_columns)
    _style_table_header(summary_sheet, summary_table_header_row)

    summary_data_start = summary_sheet.max_row + 1
    for row in summary_rows:
        summary_sheet.append([row[column] for column in summary_columns])
    summary_data_end = summary_sheet.max_row
    _apply_currency_format(
        summary_sheet, summary_data_start, summary_data_end, range(2, len(summary_columns) + 1)
    )
    if summary_data_start <= summary_data_end:
        for cell in summary_sheet[summary_data_start]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor="F7E4D5")

    storno_title_row = summary_sheet.max_row + 2
    summary_sheet.cell(
        row=storno_title_row,
        column=1,
        value="Stornierungen / Gutschriften nach Betrieb (Stückpreis × stornierte Menge)",
    )
    summary_sheet.cell(row=storno_title_row, column=1).font = Font(bold=True, size=12)
    storno_header_row = storno_title_row + 1
    storno_columns = [
        "Scope",
        "Anzahl Stornierungen",
        "Storno Brutto",
        "Storno Netto",
        "Storno Mehrwertsteuer 8.1%",
        "Storno Mehrwertsteuer 2.6%",
        "Storno Umsatz 8.1 % (Brutto)",
        "Storno Umsatz 2.6 % (Brutto)",
    ]
    summary_sheet.append(storno_columns)
    _style_table_header(summary_sheet, storno_header_row, fill_color="9B2D20")
    storno_data_start = summary_sheet.max_row + 1
    for row in storno_rows:
        summary_sheet.append([row[column] for column in storno_columns])
    storno_data_end = summary_sheet.max_row
    _apply_currency_format(summary_sheet, storno_data_start, storno_data_end, range(3, len(storno_columns) + 1))
    if storno_data_start <= storno_data_end:
        for cell in summary_sheet[storno_data_start]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor="F3D9D5")

    payment_title_row = summary_sheet.max_row + 2
    summary_sheet.cell(row=payment_title_row, column=1, value="Zahlungsmethoden nach Betrieb")
    summary_sheet.cell(row=payment_title_row, column=1).font = Font(bold=True, size=12)
    payment_header_row = payment_title_row + 1
    summary_sheet.append(["Scope", "PaymentMethod", "Total"])
    _style_table_header(summary_sheet, payment_header_row, fill_color="6B6152")
    payment_data_start = summary_sheet.max_row + 1
    for row in payment_rows:
        summary_sheet.append([row["Scope"], row["PaymentMethod"], row["Total Zahlungsmethode"]])
    payment_data_end = summary_sheet.max_row
    _apply_currency_format(summary_sheet, payment_data_start, payment_data_end, [3])

    details_sheet = workbook.create_sheet("Details")
    detail_columns = [
        "BranchName",
        "BranchUUID",
        "ItemId",
        "CreateDate",
        "OrderNumber",
        "OrderType",
        "DeviceType",
        "OrderStatus",
        "PaymentMethod",
        "PaymentReferenceId",
        "TotalAmount",
        "SubTotal",
        "DiscountAmount",
        "TaxAmount",
        "TipAmount",
        "DeliveryCost",
        "ProductName",
        "ProductId",
        "ProductVariationName",
        "CategoryName",
        "Quantity",
        "UnitPrice",
        "LineDiscountAmount",
        "ModifierTotal",
        "TaxRate",
        "LineGrossRevenue",
        "LineNetRevenue",
        "LineTaxAmount",
        "RefundedQuantity",
        "RefundGrossAmount",
        "RefundNetAmount",
        "RefundTaxAmount",
        "RefundedReason",
        "RefundedTime",
    ]
    details_sheet.append(detail_columns)
    _style_table_header(details_sheet, 1, fill_color="1E1B16")
    for row in detail_rows:
        details_sheet.append([row.get(column) for column in detail_columns])
    details_sheet.freeze_panes = "A2"
    details_sheet.auto_filter.ref = details_sheet.dimensions
    if detail_rows:
        money_detail_columns = [
            detail_columns.index(name) + 1
            for name in (
                "TotalAmount",
                "SubTotal",
                "DiscountAmount",
                "TaxAmount",
                "TipAmount",
                "Quantity",
                "UnitPrice",
                "ModifierTotal",
                "TaxRate",
                "LineGrossRevenue",
                "LineNetRevenue",
                "LineTaxAmount",
                "RefundedQuantity",
                "RefundGrossAmount",
                "RefundNetAmount",
                "RefundTaxAmount",
            )
        ]
        _apply_currency_format(details_sheet, 2, details_sheet.max_row, money_detail_columns)

    for sheet in (summary_sheet, details_sheet):
        for index, column_cells in enumerate(sheet.columns, start=1):
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(index)].width = min(max_length + 2, 40)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue(), errors, events


# ─────────────────────────────────────────────────────────────────────────────
#  Burgermeister – Order Export (one sheet per branch)
# ─────────────────────────────────────────────────────────────────────────────

def build_burgermeister_export(req: RunRequest) -> Tuple[bytes, List[str], List[ProgressEvent]]:
    """
    Export Burgermeister orders as XLSX with one sheet per branch.
    Columns: CreateDate, OrderNumber, TotalAmount, SubTotal, DiscountAmount,
             TaxAmount, TipAmount, PaymentMethod, OrderType.
    """
    orders, events, errors = fetch_accounting_orders(req)
    any_ok = any(event.status == "ok" for event in events)
    if not orders and errors and not any_ok:
        raise RuntimeError("Export failed – no rows collected. Errors: " + "; ".join(errors[:5]))

    orders = _deduplicate_retry_orders(orders)

    # Group by branch name (canonical order from BURGERMEISTER_BRANCHES)
    branch_order: Dict[str, List[Dict[str, Any]]] = {
        name: [] for name in BURGERMEISTER_BRANCHES.values()
    }
    for order in orders:
        branch_id = str(order.get("_queried_branch") or order.get("BranchUUID") or "")
        branch_name = _branch_label(branch_id)
        if branch_name in branch_order:
            branch_order[branch_name].append(order)

    HEADER = [
        "CreateDate", "OrderNumber", "TotalAmount", "SubTotal",
        "DiscountAmount", "TaxAmount", "TipAmount", "PaymentMethod", "OrderType",
    ]
    FILL_HDR = PatternFill(fill_type="solid", fgColor="CC0000")
    FONT_HDR = Font(bold=True, color="FFFFFF", size=10)
    FILL_ALT = PatternFill(fill_type="solid", fgColor="FFF5F5")

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for branch_name, branch_orders in branch_order.items():
        ws = wb.create_sheet(title=branch_name[:31])
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        # Header row
        for col, hdr in enumerate(HEADER, start=1):
            cell = ws.cell(row=1, column=col, value=hdr)
            cell.fill = FILL_HDR
            cell.font = FONT_HDR
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 18
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADER))}1"

        # Sort by CreateDate ascending
        sorted_orders = sorted(branch_orders, key=lambda o: str(o.get("CreateDate") or ""))

        for row_idx, order in enumerate(sorted_orders, start=2):
            # Parse CreateDate
            raw_date = str(order.get("CreateDate") or "")
            try:
                dt = datetime.fromisoformat(raw_date.replace("Z", "+00:00"))
                create_date_str = dt.strftime("%d.%m.%Y %H:%M")
            except Exception:
                create_date_str = raw_date

            row_values = [
                create_date_str,
                order.get("OrderNumber"),
                _money(_decimal(order.get("TotalAmount"))),
                _money(_decimal(order.get("SubTotal"))),
                _money(_decimal(order.get("DiscountAmount"))),
                _money(_decimal(order.get("TaxAmount"))),
                _money(_decimal(order.get("TipAmount"))),
                order.get("PaymentMethod"),
                order.get("OrderType"),
            ]
            bg = FILL_ALT if row_idx % 2 == 0 else None
            for col, val in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.font = Font(size=9)
                if bg:
                    cell.fill = bg
                if col in (3, 4, 5, 6, 7):
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0.00'

        # Auto column widths
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 30)

    if not wb.sheetnames:
        ws = wb.create_sheet("Keine Daten")
        ws.cell(row=1, column=1, value="Keine Bestellungen im gewählten Zeitraum.")

    bm_output = BytesIO()
    wb.save(bm_output)
    return bm_output.getvalue(), errors, events
