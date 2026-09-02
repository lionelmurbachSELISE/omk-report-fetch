from __future__ import annotations

import csv
import io
from collections import defaultdict
from io import BytesIO
from typing import Any, Callable, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def to_csv_bytes(columns: List[str], rows: List[Dict[str, Any]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(columns)
    for row in rows:
        writer.writerow([row.get(col) for col in columns])
    return buffer.getvalue().encode("utf-8")


def _safe_sheet_name(name: str) -> str:
    for ch in r"\/?*[]:'":
        name = name.replace(ch, "-")
    return name[:31]


def to_xlsx_bytes_by_branch(
    columns: List[str],
    rows: List[Dict[str, Any]],
    branch_label_fn: Optional[Callable[[str], str]] = None,
) -> bytes:
    by_branch: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        branch_id = str(row.get("BranchUUID") or "Unknown")
        by_branch[branch_id].append(row)

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(fill_type="solid", fgColor="1E1B16")
    header_font = Font(bold=True, color="FFFFFF")

    if not by_branch:
        ws = wb.create_sheet("No Data")
        ws.append(columns)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
    else:
        for branch_id, branch_rows in by_branch.items():
            label = branch_label_fn(branch_id) if branch_label_fn else branch_id
            ws = wb.create_sheet(_safe_sheet_name(label))
            ws.append(columns)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
            for row in branch_rows:
                ws.append([row.get(col) for col in columns])
            for idx, col_cells in enumerate(ws.columns, 1):
                max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
                ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 40)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
